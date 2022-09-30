# RT_reporter.py - Read the work.xslx file, verify it and input the time spent
# into the specified RT ticket.
# Usage: phtyon RT_reporter.py location_of_work_excel

import getpass
import sys
import re

from openpyxl import load_workbook

import RequestTracker
from helpers import exit_script

RT_SERVER = 'https://rt.cosylab.com/REST/1.0/'

if len(sys.argv) != 2:
    print('Usage: python RT_reporter.py location/of/work.xslx')
    exit_script()

print('Welcome to the RT Reporter!')

# Open the excel workbook, based on the filename passed.
# open the sheet twice, once with the formulas converted to data, in order to 
# use it and once as the original sheet, in order to update the T_input column.
filename = sys.argv[1]
try:
    wb_data = load_workbook(filename=filename, data_only=True)
    wb = load_workbook(filename=filename)
except PermissionError:
    print('Cannot open the .xlsx file, close before proceeding.')
    exit_script()
except FileNotFoundError:
    print('The .xlsx file cannot be found at the specified location.')
    exit_script()

# Ask for the username and password.
username = input('Please input your username: ')
password = getpass.getpass('Please input your password: ')
print('')

# Log into RT.
rt = RequestTracker.RequestTracker(RT_SERVER)
rt.login(username, password)

# Column identifiers of the significatnt columns in the project sheets.
TASK_COLUMN = 1
T_SPENT_COLUMN = 2
RT_LINK_COLUMN = 4
T_INPUT_COLUMN = 5

# Prepare the regex for extracting the ticket ID number from the RT link.
rt_link_regex = re.compile(f'\?id=(\d+)$')
reply_text_regex = re.compile(f'(.) *(.*)')

# Before inputting the time, check that all the time in the 'Time' sheet has 
# been accounted for in the project sheets.
sheet_data = wb_data['Analysis']
total_time = sheet_data['B3'].value
accounted_time = sheet_data['B7'].value
if accounted_time and total_time != accounted_time:
    print('Not all of the input time has been accounted for in the project sheets!')
    print('Fix the issue and re-run the script.')
    exit_script()

# Loop through all the project sheets in the excel.
for sheetname in wb.sheetnames[1:-1]:
    sheet_data = wb_data[sheetname]
    sheet = wb[sheetname]
    project_text = f"| Reporting time for project: {sheet_data.title} |"
    print('-' * len(project_text))
    print(project_text)
    print('-' * len(project_text), end = '\n\n')

    # Loop through the tasks in a project sheet.
    for row in range(5, sheet_data.max_row + 1):
        task = sheet_data.cell(row=row, column=TASK_COLUMN).value

        # If no value has been input into the spent time, set it to zero.
        # This handles blank rows in the sheet, if any.
        t_spent = sheet_data.cell(row=row, column=T_SPENT_COLUMN).value
        t_spent = (0 if t_spent==None else round(t_spent))

        # If no value has been input into the input time, set it to zero
        t_input = sheet_data.cell(row=row, column=T_INPUT_COLUMN).value
        t_input = (0 if t_input==None else t_input)

        rt_link = sheet_data.cell(row=row, column=RT_LINK_COLUMN).value

        if (t_for_rt := t_spent-t_input) > 0:
            print(f'Inputting time for {task}')
            if rt_link:
                try:
                    mo = rt_link_regex.search(str(rt_link))
                    if mo:
                        rt_number = int(mo[1])
                    else:
                        # If the complete rt link is not input, try to see if the user input the ticket number only.
                        rt_number = int(rt_link)
                    rt_name = rt.get_name(rt_number)
                except ValueError:
                    print(f"Invalid ticket number for task '{task}' in RT link: {rt_link}.")
                    print('')
                    continue
                except TypeError:
                    print(f"Cannot find a valid ticket with number {rt_link} for task '{task}', skipping.")
                    print('')
                    continue

                print(f"Task: {task}, RT#{rt_number}: {rt_name}, Time to input: {t_for_rt}.")
                response = input("'c [Text for reply.]' to confirm. ")

                # Find the obligatory 'c', followed by the optional reply text.
                mo = reply_text_regex.search(response) 
                try:
                    if mo[1] == 'c':
                        print(f"Updating RT#{rt_number} with {t_for_rt} minutes.")
                        if mo[2]:
                            rt.reply(rt_number, mo[2], t_for_rt)
                        else:
                            rt.reply(rt_number, 't', t_for_rt)
                        sheet.cell(row=row, column=T_INPUT_COLUMN).value = t_for_rt + t_input
                except TypeError:    
                    None

            else:
                print(f"Task '{task}' does not have a ticket defined, skipping.")

            print('')
            
# Save the work excel sheet with the values, input into RT.
wb.save(filename=filename)
rt.logout()

print('Finished reporting to RT.')
exit_script()